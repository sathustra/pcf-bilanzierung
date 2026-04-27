from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
import io
import models
from database import get_db
from calculations import calculate_pcf

router = APIRouter(prefix="/api/calculate", tags=["calculations"])


def _load_product_data(product_id: int, db: Session):
    product = db.query(models.Product).filter(models.Product.id == product_id).first()
    if not product:
        raise HTTPException(status_code=404, detail="Produkt nicht gefunden")

    materials = (db.query(models.MaterialEntry)
                 .options(joinedload(models.MaterialEntry.emission_factor))
                 .filter(models.MaterialEntry.product_id == product_id).all())
    energy = (db.query(models.EnergyEntry)
              .options(joinedload(models.EnergyEntry.emission_factor))
              .filter(models.EnergyEntry.product_id == product_id).all())
    transport = (db.query(models.TransportEntry)
                 .options(joinedload(models.TransportEntry.emission_factor))
                 .filter(models.TransportEntry.product_id == product_id).all())
    waste = (db.query(models.WasteEntry)
             .options(joinedload(models.WasteEntry.emission_factor))
             .filter(models.WasteEntry.product_id == product_id).all())
    return product, materials, energy, transport, waste


@router.get("/{product_id}")
def get_pcf_result(product_id: int, db: Session = Depends(get_db)):
    product, materials, energy, transport, waste = _load_product_data(product_id, db)
    result = calculate_pcf(product, materials, energy, transport, waste)
    return {
        "product_id": product_id,
        "product_name": product.product_name,
        "reference_unit": product.reference_unit,
        "annual_production": product.annual_production,
        **result,
    }


@router.get("/{product_id}/export/excel")
def export_excel(product_id: int, db: Session = Depends(get_db)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    product, materials, energy, transport, waste = _load_product_data(product_id, db)
    result = calculate_pcf(product, materials, energy, transport, waste)

    wb = Workbook()
    ws = wb.active
    ws.title = "PCF Ergebnis"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2D6A4F")

    def write_header(row, cols):
        for col_idx, col_name in enumerate(cols, 1):
            cell = ws.cell(row=row, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

    # Title
    ws["A1"] = f"PCF-Bilanz: {product.product_name}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Bezugseinheit: {product.reference_unit} | Jahresproduktion: {product.annual_production}"
    ws["A3"] = f"Referenzjahr: {product.reference_year} | Scope: {product.scope}"

    # GWP Indicators
    row = 5
    ws.cell(row=row, column=1, value="GWP-Indikatoren (kg CO₂eq / Bezugseinheit)")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
    row += 1
    write_header(row, ["Indikator", "Wert"])
    row += 1
    gwp = result["gwp_indicators"]
    for name, val in [
        ("GWP-fossil", gwp["gwp_fossil"]),
        ("GWP-biogenic emissions", gwp["gwp_biogenic_emissions"]),
        ("GWP-biogenic uptake", gwp["gwp_biogenic_uptake"]),
        ("GWP-LULUC", gwp["gwp_luluc"]),
        ("GWP-total", gwp["gwp_total"]),
    ]:
        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=2, value=val)
        if name == "GWP-total":
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.cell(row=row, column=2).font = Font(bold=True)
        row += 1

    # Life Cycle Stages
    row += 1
    ws.cell(row=row, column=1, value="Life Cycle Stages")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
    row += 1
    write_header(row, ["Stage", "Biogen", "LUC", "Fossil", "Anteil GWP-total %", "Anteil GWP-fossil %"])
    row += 1
    for lcs in result["lifecycle_results"]:
        ws.cell(row=row, column=1, value=lcs["stage"])
        ws.cell(row=row, column=2, value=lcs["biogenic"])
        ws.cell(row=row, column=3, value=lcs["luc"])
        ws.cell(row=row, column=4, value=lcs["fossil"])
        ws.cell(row=row, column=5, value=lcs["share_of_gwp_total"])
        ws.cell(row=row, column=6, value=lcs["share_of_gwp_fossil"])
        row += 1

    # Sub-categories
    row += 1
    ws.cell(row=row, column=1, value="Sub-Kategorien")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
    row += 1
    write_header(row, ["Sub-Kategorie", "Life Cycle Stage", "Biogen", "LUC", "Fossil"])
    row += 1
    for sc in result["subcategory_results"]:
        ws.cell(row=row, column=1, value=sc["sub_category"])
        ws.cell(row=row, column=2, value=sc["life_cycle_stage"])
        ws.cell(row=row, column=3, value=sc["biogenic"])
        ws.cell(row=row, column=4, value=sc["luc"])
        ws.cell(row=row, column=5, value=sc["fossil"])
        row += 1

    # DQR
    row += 1
    ws.cell(row=row, column=1, value="Datenqualität (DQR)")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
    row += 1
    dqr = result["dqr"]
    ws.cell(row=row, column=1, value="DQR-Score")
    ws.cell(row=row, column=2, value=dqr["dqr_score"])
    row += 1
    ws.cell(row=row, column=1, value="Ampel")
    ws.cell(row=row, column=2, value=dqr["traffic_light"])
    row += 1
    ws.cell(row=row, column=1, value="Primärdatenanteil %")
    ws.cell(row=row, column=2, value=dqr["primary_data_share"])

    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 40

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"PCF_{product.product_name.replace(' ', '_')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
